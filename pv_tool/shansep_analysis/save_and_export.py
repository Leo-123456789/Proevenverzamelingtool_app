from typing import TYPE_CHECKING, List
from pandas import ExcelWriter, concat, DataFrame, read_excel
from datetime import datetime
from pv_tool.utilities.utils import get_repo_root
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, LongTable
from pv_tool.imports.excel_utils import format_excel_sheet
import plotly.graph_objects as go
import numpy as np
from pathlib import Path
from reportlab.platypus import PageBreak, Image as RLImage

try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    PILImage = None
    print("Waarschuwing: PIL (Pillow) is niet beschikbaar. Figuren kunnen niet aan PDF worden toegevoegd.")

if TYPE_CHECKING:
    from pv_tool.shansep_analysis.shansep_analysis import SHANSEP


# def add_results_to_dbase(self: "SHANSEP", path: str,
#                          file_name: str = 'Template_PVtool5_0.xlsx'):  # TODO deze wordt vervangen door add_results_to_template
#     """
#     Voegt de SHANSEP analyseresultaten toe aan de database Excel-bestand.
#
#     Parameters
#     ----------
#     self : SHANSEP
#         Instantie van de SHANSEP analyse klasse
#     path : str
#         Pad naar de map waar het Excel-bestand staat
#     file_name : str
#         Naam van het Excel-bestand
#
#     Returns
#     -------
#     DataFrame
#         Bijgewerkte DataFrame met alle resultaten
#     """
#     file_path = f"{path}/{file_name}"
#
#     # Run analysis to get results
#     df_gem, df_kar = self.get_result_values_shansep()
#
#     # Expected columns structure voor SHANSEP resultaten
#     expected_columns = [
#         'PVNAAM', 'PV_REK', 'PV_TYPE_PROEF', 'PV_ANALYSE', 'PV_RESULTAAT_ID', 'PV_TYPEVERZAMELING',
#         'PV_A1_SNIJPUNT_YAS_GEM [-]', 'PV_A2_S_GEM [-]', 'PV_m_GEM [-]', 'PV_POP_GEM [kPa]',
#         'PV_A1_SNIJPUNT_YAS_KAR [-]', 'PV_A2_S_KAR [-]', 'PV_m_KAR [-]', 'PV_POP_KAR [kPa]',
#         'PV_S_SD_DSTAB [-]', 'PV_m_SD_DSTAB [-]', 'PV_POP_SD_DSTAB [-]',
#         'PV_VGWNAT_GEM [kN/m3]', 'PV_VGWNAT_SD [kN/m3]', 'PV_WATERGEHALTE_GEM', 'PV_WATERGEHALTE_SD',
#         'Timestamp'
#     ]
#
#     new_row = {
#         'PVNAAM': self.investigation_groups[0],
#         'PV_REK': self.effective_stress,
#         'PV_TYPE_PROEF': self.analysis_type.split('_')[0],
#         'PV_ANALYSE': '_'.join(self.analysis_type.split('_')[1:]),
#         'PV_RESULTAAT_ID': f"{self.investigation_groups[0]}_{self.effective_stress}_{self.analysis_type}",
#         'PV_TYPEVERZAMELING': self.alpha,
#         'PV_A1_SNIJPUNT_YAS_GEM [-]': round(self.snijpunt_gem_handmatig,
#                                             3) if self.snijpunt_gem_handmatig is not None else None,
#         'PV_A2_S_GEM [-]': round(self.s_gem_handmatig, 3) if self.s_gem_handmatig is not None else None,
#         'PV_m_GEM [-]': round(self.m_gem_handmatig, 3) if self.m_gem_handmatig is not None else None,
#         'PV_POP_GEM [kPa]': round(self.pop_gem_handmatig, 3) if self.pop_gem_handmatig is not None else None,
#         'PV_A1_SNIJPUNT_YAS_KAR [-]': round(self.snijpunt_kar_handmatig,
#                                             3) if self.snijpunt_kar_handmatig is not None else None,
#         'PV_A2_S_KAR [-]': round(self.s_kar_handmatig, 3) if self.s_kar_handmatig is not None else None,
#         'PV_m_KAR [-]': round(self.m_kar_handmatig, 3) if self.m_kar_handmatig is not None else None,
#         'PV_POP_KAR [kPa]': round(self.pop_kar_handmatig, 3) if self.pop_kar_handmatig is not None else None,
#         'PV_S_SD_DSTAB [-]': round(self.st_dev_s_handmatig, 3) if self.st_dev_s_handmatig is not None else None,
#         'PV_m_SD_DSTAB [-]': round(self.st_dev_m_handmatig, 3) if self.st_dev_s_handmatig is not None else None,
#         'PV_POP_SD_DSTAB [-]': round(self.st_dev_pop_handmatig, 3) if self.st_dev_pop_handmatig is not None else None,
#         'PV_VGWNAT_GEM [kN/m3]': round(self.calc_vgwnat_gem, 3) if self.calc_vgwnat_gem is not None else None,
#         'PV_VGWNAT_SD [kN/m3]': round(self.calc_vgwnat_sd, 3) if self.calc_vgwnat_sd is not None else None,
#         'PV_WATERGEHALTE_GEM': round(self.calc_watergehalte_gem, 3) if self.calc_watergehalte_gem is not None else None,
#         'PV_WATERGEHALTE_SD': round(self.calc_watergehalte_sd, 3) if self.calc_watergehalte_sd is not None else None,
#         'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#     }
#
#     workbook = load_workbook(file_path)
#
#     if 'Resultaten SHANSEP' in workbook.sheetnames:
#         print('Tabblad resultaten SHANSEP in dbase excel bestaat al en wordt aangevuld')
#         df_existing = read_excel(file_path, sheet_name='Resultaten SHANSEP')
#         # Filter out empty rows and ensure consistent types before concatenation
#         df_existing = df_existing.dropna(how='all')
#         # Ensure column headers are strings
#         df_existing.columns = df_existing.columns.astype(str)
#         new_row_df = DataFrame([new_row], columns=df_existing.columns)
#         df_updated = concat([df_existing, new_row_df], ignore_index=True)
#     else:
#         print('Tabblad resultaten SHANSEP in dbase excel bestaat nog niet en wordt aangemaakt')
#         df_updated = DataFrame([new_row], columns=expected_columns)
#
#     # Ensure all column headers are strings
#     df_updated.columns = df_updated.columns.astype(str)
#
#     # Write data to Excel
#     with ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#         df_updated.to_excel(writer, sheet_name='Resultaten SHANSEP', index=False)
#
#     # Formatting
#     num_columns = df_updated.shape[1]
#     num_rows = df_updated.shape[0]
#     format_excel_sheet(
#         file_path=file_path,
#         sheet_name='Resultaten SHANSEP',
#         num_columns=num_columns,
#         num_rows=num_rows,
#         table_name='ResultatenSHANSEPTable',
#         index=False
#     )
#
#     return df_updated


def add_results_to_template(self: "SHANSEP", path, export_name=None):
    """
    Voegt de SHANSEP analyseresultaten toe aan de database Excel-bestand.

    Parameters
    ----------
    self: SHANSEP
        Instantie van de SHANSEP analyse klasse
    path
        Pad naar de map waar het Excel-bestand staat
    export_name
        Naam van het Excel-bestand

    Returns
    -------
    DataFrame
        Bijgewerkte DataFrame met alle resultaten
    """
    if export_name is None:
        export_name = "Template_PVtool5_0.xlsx"
    file_path = Path(path) / export_name
    sheet_name = 'Resultaten SHANSEP'

    # Run analysis to get results
    _, _ = self.get_result_values_shansep()

    # Expected columns structure voor SHANSEP resultaten
    expected_columns = [
        'PVNAAM', 'PV_REK', 'PV_TYPE_PROEF', 'PV_ANALYSE', 'PV_RESULTAAT_ID', 'PV_TYPEVERZAMELING',
        'PV_A1_SNIJPUNT_YAS_GEM [-]', 'PV_A2_S_GEM [-]', 'PV_m_GEM [-]', 'PV_POP_GEM [kPa]',
        'PV_A1_SNIJPUNT_YAS_KAR [-]', 'PV_A2_S_KAR [-]', 'PV_m_KAR [-]', 'PV_POP_KAR [kPa]',
        'PV_S_SD_DSTAB [-]', 'PV_m_SD_DSTAB [-]', 'PV_POP_SD_DSTAB [-]',
        'PV_VGWNAT_GEM [kN/m3]', 'PV_VGWNAT_SD [kN/m3]', 'PV_WATERGEHALTE_GEM', 'PV_WATERGEHALTE_SD',
        'Timestamp'
    ]

    new_row = {
        'PVNAAM': self.investigation_groups[0],
        'PV_REK': self.effective_stress,
        'PV_TYPE_PROEF': self.analysis_type.split('_')[0],
        'PV_ANALYSE': '_'.join(self.analysis_type.split('_')[1:]),
        'PV_RESULTAAT_ID': f"{self.investigation_groups[0]}_{self.effective_stress}_{self.analysis_type}",
        'PV_TYPEVERZAMELING': self.alpha,
        'PV_A1_SNIJPUNT_YAS_GEM [-]': round(self.snijpunt_gem_handmatig,
                                            3) if self.snijpunt_gem_handmatig is not None else None,
        'PV_A2_S_GEM [-]': round(self.s_gem_handmatig, 3) if self.s_gem_handmatig is not None else None,
        'PV_m_GEM [-]': round(self.m_gem_handmatig, 3) if self.m_gem_handmatig is not None else None,
        'PV_POP_GEM [kPa]': round(self.pop_gem_handmatig, 3) if self.pop_gem_handmatig is not None else None,
        'PV_A1_SNIJPUNT_YAS_KAR [-]': round(self.snijpunt_kar_handmatig,
                                            3) if self.snijpunt_kar_handmatig is not None else None,
        'PV_A2_S_KAR [-]': round(self.s_kar_handmatig, 3) if self.s_kar_handmatig is not None else None,
        'PV_m_KAR [-]': round(self.m_kar_handmatig, 3) if self.m_kar_handmatig is not None else None,
        'PV_POP_KAR [kPa]': round(self.pop_kar_handmatig, 3) if self.pop_kar_handmatig is not None else None,
        'PV_S_SD_DSTAB [-]': round(self.st_dev_s_handmatig, 3) if self.st_dev_s_handmatig is not None else None,
        'PV_m_SD_DSTAB [-]': round(self.st_dev_m_handmatig, 3) if self.st_dev_s_handmatig is not None else None,
        'PV_POP_SD_DSTAB [-]': round(self.st_dev_pop_handmatig, 3) if self.st_dev_pop_handmatig is not None else None,
        'PV_VGWNAT_GEM [kN/m3]': round(self.calc_vgwnat_gem, 3) if self.calc_vgwnat_gem is not None else None,
        'PV_VGWNAT_SD [kN/m3]': round(self.calc_vgwnat_sd, 3) if self.calc_vgwnat_sd is not None else None,
        'PV_WATERGEHALTE_GEM': round(self.calc_watergehalte_gem, 3) if self.calc_watergehalte_gem is not None else None,
        'PV_WATERGEHALTE_SD': round(self.calc_watergehalte_sd, 3) if self.calc_watergehalte_sd is not None else None,
        'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    workbook = load_workbook(file_path)

    if 'Resultaten SHANSEP' in workbook.sheetnames:
        print('Tabblad resultaten SHANSEP in dbase excel bestaat al en wordt aangevuld')
        df_existing = read_excel(file_path, sheet_name='Resultaten SHANSEP')
        # Filter out empty rows and ensure consistent types before concatenation
        df_existing = df_existing.dropna(how='all')
        # Ensure column headers are strings
        df_existing.columns = df_existing.columns.astype(str)
        new_row_df = DataFrame([new_row], columns=df_existing.columns)
        df_updated = concat([df_existing, new_row_df], ignore_index=True)
    else:
        print('Tabblad resultaten SHANSEP in dbase excel bestaat nog niet en wordt aangemaakt')
        df_updated = DataFrame([new_row], columns=expected_columns)

    # Ensure all column headers are strings
    df_updated.columns = df_updated.columns.astype(str)

    if file_path.exists():
        wb = load_workbook(file_path)
    else:
        template_path = Path(get_repo_root()) / "pv_tool" / "templates" / "Template_PVtool5_0.xlsx"
        wb = load_workbook(template_path)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        first_empty_row = ws.max_row + 1 if any(
            ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True)) else ws.max_row
    else:
        ws = wb.create_sheet(sheet_name)
        for col_idx, col_name in enumerate(expected_columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        first_empty_row = 2

    for col_idx, col_name in enumerate(expected_columns, start=1):
        ws.cell(row=first_empty_row, column=col_idx, value=new_row.get(col_name, ""))

    wb.save(file_path)
    print(f"Resultaat toegevoegd aan template in tabblad '{sheet_name}'.")


def save_total_to_excel(self: "SHANSEP", path: str):
    """
    Exporteert alle SHANSEP analysegegevens naar Excel.

    Slaat de volledige dataset met alle berekende kolommen op in een Excel bestand.
    De bestandsnaam wordt automatisch gegenereerd op basis van de analyse-instellingen.

    Parameters
    ----------
    path: str
        Map locatie waar het Excel-bestand moet worden opgeslagen
    self: SHANSEP
        Instantie van de SHANSEP klasse
    """

    # Pas de effective stress naam aan zodat het weggeschreven kan worden in de bestandsnaam
    effective_stress = str(self.effective_stress).replace('%', 'procent_')
    effective_stress = str(effective_stress).replace(' ', '')

    # Exporteer onder de juiste naam
    file_name = f"shansep_export_{self.investigation_groups[0]}_{self.analysis_type}_{effective_stress}.xlsx"
    file_path = f"{path}/{file_name}"

    # Run analysis to ensure data is available
    self._run_shansep()

    # Get results
    df_gem, df_kar = self.get_result_values_shansep()

    # Ensure all column headers are strings
    if self.shansep_data_df_oc is not None:
        self.shansep_data_df_oc.columns = self.shansep_data_df_oc.columns.astype(str)
    if self.shansep_data_df_nc_oc is not None:
        self.shansep_data_df_nc_oc.columns = self.shansep_data_df_nc_oc.columns.astype(str)
    if hasattr(self, 'sutabel') and self.sutabel is not None:
        self.sutabel.columns = self.sutabel.columns.astype(str)

    # Ensure result DataFrames have string columns and index
    df_gem.columns = df_gem.columns.astype(str)
    df_kar.columns = df_kar.columns.astype(str)

    # Ensure index is also string type
    df_gem.index = df_gem.index.astype(str)
    df_kar.index = df_kar.index.astype(str)

    # Set index names to avoid Excel adding 'Column1' header
    df_gem.index.name = 'analyse'
    df_kar.index.name = 'analyse'

    # Write all data to Excel
    with ExcelWriter(file_path, engine='openpyxl') as writer:
        # Main analysis data
        if self.shansep_data_df_oc is not None:
            self.shansep_data_df_oc.to_excel(writer, sheet_name='Analyse Data OC', index=False)
        if self.shansep_data_df_nc_oc is not None:
            self.shansep_data_df_nc_oc.to_excel(writer, sheet_name='Analyse Data OC en NC', index=False)

        # Results - clean data before export
        df_gem_export = df_gem.copy()
        df_kar_export = df_kar.copy()

        # Replace None values with empty string to avoid Excel issues
        df_gem_export = df_gem_export.fillna('')
        df_kar_export = df_kar_export.fillna('')

        df_gem_export.to_excel(writer, sheet_name='Resultaten Gemiddeld', index=True)
        df_kar_export.to_excel(writer, sheet_name='Resultaten Karakteristiek', index=True)

        # Su tabel if available
        if hasattr(self, 'sutabel') and self.sutabel is not None:
            self.sutabel.to_excel(writer, sheet_name='Su Tabel', index=False)

    # format the sheets
    if self.shansep_data_df_oc is not None:
        format_excel_sheet(
            file_path=file_path,
            sheet_name='Analyse Data OC',
            num_columns=self.shansep_data_df_oc.shape[1],
            num_rows=self.shansep_data_df_oc.shape[0],
            table_name='AnalyseDataOCTable',
            index=False
        )
    if self.shansep_data_df_nc_oc is not None:
        format_excel_sheet(
            file_path=file_path,
            sheet_name='Analyse Data OC en NC',
            num_columns=self.shansep_data_df_nc_oc.shape[1],
            num_rows=self.shansep_data_df_nc_oc.shape[0],
            table_name='AnalyseDataNC_OCTable',
            index=False
        )
    format_excel_sheet(
        file_path=file_path,
        sheet_name='Resultaten Gemiddeld',
        num_columns=df_gem.shape[1],  # +1 for index
        num_rows=df_gem.shape[0],
        table_name='ResultatenGemiddeldTable',
        index=True
    )
    format_excel_sheet(
        file_path=file_path,
        sheet_name='Resultaten Karakteristiek',
        num_columns=df_kar.shape[1],  # +1 for index
        num_rows=df_kar.shape[0],
        table_name='ResultatenKarakteristiekTable',
        index=True
    )
    if self.sutabel is not None:
        format_excel_sheet(
            file_path=file_path,
            sheet_name='Su Tabel',
            num_columns=self.sutabel.shape[1],
            num_rows=self.sutabel.shape[0],
            table_name='SuTabelTable',
            index=False
        )

    print(f"SHANSEP Excel export voltooid: {file_path}")


def _df_to_table_with_index(df, index_name='Index'):
    """
    Zet een DataFrame om naar een lijst voor gebruik in een PDF-tabel.

    Parameters
    ----------
    df: DataFrame
        De DataFrame die moet worden omgezet
    index_name: str, optioneel
        Naam voor de index kolom (default='Index')

    Returns
    -------
    list
        Lijst met header en data rijen voor een PDF-tabel
    """
    header = [df.index.name or index_name] + df.columns.tolist()
    data = [[idx] + row.tolist() for idx, row in df.iterrows()]
    return [header] + data


def _create_input_table(self: "SHANSEP") -> Table:
    """
    Maakt een tabel met de invoerselectie informatie voor SHANSEP analyse.

    Returns
    -------
    Table
        ReportLab tabel object met de invoerselectie informatie
    """
    if self.shansep_data_df_nc_oc_unsorted is None:
        return Table([['Geen invoerdata beschikbaar']], hAlign='LEFT')

    # Selecteer relevante kolommen
    columns_base = ['PV_NAAM', 'BORING_POSITIE', 'MONSTER_NIVEAU_NAP_VANAF', 'MONSTER_NIVEAU_NAP_TOT',
                    'ANA_TERREINSPANNING', 'ANA_GRENSSPANNING_REKEN', 'ANA_POP_VELD']

    if self.analysis_type.startswith('TXT'):
        columns_extra = ['TXT_SS_VOLUMEGEWICHT_NAT', 'TXT_SS_VOLUMEGEWICHT_DRG', 'TXT_SS_WATERGEHALTE_VOOR']
    else:
        columns_extra = ['DSS_VOLUMEGEWICHT_NAT', 'DSS_VOLUMEGEWICHT_DRG', 'DSS_WATERGEHALTE_VOOR']

    # Check welke kolommen bestaan in het dataframe
    available_columns = [col for col in columns_base + columns_extra if col in self.total_shansep_data_df.columns]

    if not available_columns:
        return Table([['Geen invoerdata kolommen beschikbaar']], hAlign='LEFT')

    table_df = self.total_shansep_data_df[available_columns].copy()

    # Add additional columns from shansep_data_df_nc_oc if available and data is properly sorted
    additional_columns = ['S\'v', 'Su', 'consolidatietype', 'OCR']
    if (hasattr(self, 'shansep_data_df_nc_oc_unsorted') and
            self.shansep_data_df_nc_oc_unsorted is not None and
            len(self.shansep_data_df_nc_oc_unsorted) == len(table_df)):

        # Check which additional columns are available
        for col in additional_columns:
            if col in self.shansep_data_df_nc_oc_unsorted.columns and col not in table_df.columns:
                # Ensure the indices match before adding the column
                if table_df.index.equals(self.shansep_data_df_nc_oc_unsorted.index):
                    table_df[col] = self.shansep_data_df_nc_oc_unsorted[col]

    else:
        print(
            f"WAARSCHUWING: extra kolommen {additional_columns} in invoer tabel konden niet worden toegevoegd aan "
            f"invoer tabel vanwege ontbrekende of niet-overeenkomende gegevens.")

    # Hernoem kolommen voor leesbaarheid met kortere namen en line breaks
    column_mapping = {
        'PV_NAAM': 'Groep',
        'BORING_POSITIE': 'Pos.',
        'MONSTER_NIVEAU_NAP_VANAF': 'NAP Van\n[m]',
        'MONSTER_NIVEAU_NAP_TOT': 'NAP Tot\n[m]',
        'TXT_SS_VOLUMEGEWICHT_NAT': 'VGW nat\n[kN/m3]',
        'TXT_SS_VOLUMEGEWICHT_DRG': 'VGW drg\n[kN/m3]',
        'TXT_SS_WATERGEHALTE_VOOR': 'Water\n[%]',
        'DSS_VOLUMEGEWICHT_NAT': 'VGW nat\n[kN/m3]',
        'DSS_VOLUMEGEWICHT_DRG': 'VGW drg\n[kN/m3]',
        'DSS_WATERGEHALTE_VOOR': 'Water\n[%]',
        'S\'v': "σ'v\n[kPa]",
        'Su': "su\n[kPa]",
        'OCR': 'OCR\n[-]',
        'ANA_TERREINSPANNING': 'Terreinspanning\n[kPa]',
        'ANA_GRENSSPANNING_REKEN': 'Grensspanning\n[kPa]',
        'ANA_POP_VELD': 'POP\n[-]'
    }

    table_df = table_df.rename(columns={k: v for k, v in column_mapping.items() if k in table_df.columns})

    # replace column POP by new calculation to account for missing values: terreinspanning*ocr-terreinspanning
    if 'Terreinspanning\n[kPa]' in table_df.columns and 'OCR\n[-]' in table_df.columns:
        table_df['POP\n[-]'] = table_df['Terreinspanning\n[kPa]'] * table_df['OCR\n[-]'] - table_df[
            'Terreinspanning\n[kPa]']

    # Sort table_df in the same way as shansep_data_df_nc_oc
    # Check if consolidatietype column exists in the data (use original column name before renaming)

    if 'consolidatietype' in table_df.columns:
        table_df = table_df.sort_values(
            by=['consolidatietype', table_df.index.name or table_df.index],
            ascending=[False, True]
        ).copy()
    else:
        table_df = table_df.sort_index()
        print(f"WAARSCHUWING: invoer tabel kon niet worden gesorteerd op 'consolidatietype' omdat de kolom ontbreekt.")

    # delete consolidatietype column after sorting
    if 'consolidatietype' in table_df.columns:
        table_df = table_df.drop(columns=['consolidatietype'])

    # Handle NaN values and format numeric values properly
    for col in table_df.columns:
        if table_df[col].dtype in ['float64', 'int64']:
            table_df[col] = table_df[col].apply(
                lambda x: f"{x:.2f}" if isinstance(x, (float, int)) and not np.isnan(x) else ""
            )
        else:
            table_df[col] = table_df[col].apply(
                lambda x: "" if (isinstance(x, float) and np.isnan(x)) or x is None else str(x)
            )

    # Create table data with proper line break handling
    # First create the header row with Paragraph objects for consistent formatting
    from reportlab.lib.styles import getSampleStyleSheet
    header_style = getSampleStyleSheet()['Normal']
    header_style.fontSize = 7  # Set to 7 as requested
    header_style.fontName = 'Helvetica-Bold'
    header_style.alignment = 0  # Left align

    # Create header row with Paragraph objects for ALL headers (consistent formatting)
    header_row = [Paragraph('Monster ID', header_style)]

    for col in table_df.columns:
        if '\n' in col:
            # Create Paragraph for multi-line headers
            header_row.append(Paragraph(col.replace('\n', '<br/>'), header_style))
        else:
            # Create Paragraph for single-line headers too (for consistency)
            header_row.append(Paragraph(col, header_style))

    # Create data rows
    data_rows = []
    for idx, row in table_df.iterrows():
        data_row = [str(idx)] + [str(val) if val != "" else "" for val in row]
        data_rows.append(data_row)

    # Combine header and data
    t1_data = [header_row] + data_rows

    # Create table with automatic column sizing
    t1 = LongTable(t1_data, repeatRows=1, hAlign='LEFT')
    t1.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Only apply font to data rows (headers are Paragraphs)
        ('FONTSIZE', (0, 1), (-1, -1), 6),  # Set data font size to 6 as requested
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),  # Header row alignment
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),  # Data row alignment
        ('LEFTPADDING', (0, 0), (-1, -1), 2),  # Reduce padding
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        # Special styling for header row with Paragraphs
        ('ROWBACKGROUNDS', (0, 0), (-1, 0), [colors.lightgrey]),
    ]))
    return t1


def _create_parameters_table(self: "SHANSEP") -> Table:
    """
    Maakt een tabel met de SHANSEP parameters.

    Returns
    -------
    Table
        ReportLab tabel object met de SHANSEP parameters
    """
    parameters: List[list] = []

    # Voeg parameters toe
    if hasattr(self, 'e_a1_oc') and self.e_a1_oc is not None and self.snijpunt_gem_handmatig is None:
        parameters.append(['Snijpunt y-as gemiddeld [kPa] (voorstel)', round(self.e_a1_oc, 3)])
    if hasattr(self, 'e_a2_oc') and self.e_a2_oc is not None and self.s_gem_handmatig is None:
        parameters.append(['S gemiddeld [-] (voorstel)', round(self.e_a2_oc, 3)])
    if hasattr(self, 'e_a2_nc_oc') and self.e_a2_nc_oc is not None and self.m_gem_handmatig is None:
        parameters.append(['m gemiddeld [-] (voorstel)', round(self.e_a2_nc_oc, 3)])
    if hasattr(self, 'pop_gem_oc') and self.pop_gem_oc is not None and self.pop_gem_handmatig is None:
        parameters.append(['POP gemiddeld [kPa] (voorstel)', round(self.pop_gem_oc, 3)])
    if hasattr(self, 'a1_kar_oc') and self.a1_kar_oc is not None and self.snijpunt_kar_handmatig is None:
        parameters.append(['Snijpunt y-as karakteristiek [kPa] (voorstel)', round(self.a1_kar_oc, 3)])
    if hasattr(self, 'a2_kar_oc') and self.a2_kar_oc is not None and self.s_kar_handmatig is None:
        parameters.append(['S karakteristiek [-] (voorstel)', round(self.a2_kar_oc, 3)])
    if hasattr(self, 'a2_kar_nc_oc') and self.a2_kar_nc_oc is not None and self.m_kar_handmatig is None:
        parameters.append(['m karakteristiek [-] (voorstel)', round(self.a2_kar_nc_oc, 3)])
    if hasattr(self, 'pop_kar_oc') and self.pop_kar_oc is not None and self.pop_kar_handmatig is None:
        parameters.append(['POP karakteristiek [kPa] (voorstel)', round(self.pop_kar_oc, 3)])

    parameters.append(['Type verzameling: lokaal = 1.0; regionaal = 0.75', self.alpha])

    if hasattr(self, 'calc_vgwnat_gem') and self.calc_vgwnat_gem is not None:
        parameters.append(['VGW nat gemiddeld [kN/m3]', round(self.calc_vgwnat_gem, 2)])
    if hasattr(self, 'calc_watergehalte_gem') and self.calc_watergehalte_gem is not None:
        parameters.append(['Watergehalte gemiddeld [%]', round(self.calc_watergehalte_gem, 2)])

    t = Table([['Parameter', 'Waarde']] + parameters, hAlign='LEFT')
    t.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    return t


def _create_gem_results_table(self: "SHANSEP") -> Table:
    """
    Maakt een tabel met de SHANSEP gemiddelde resultaten (alleen voorgestelde parameters).

    Returns
    -------
    Table
        ReportLab tabel object met de SHANSEP gemiddelde resultaten
    """
    df_gem, df_kar = self.get_result_values_shansep()

    # Filter out handmatige rows - alleen voorgestelde parameters behouden
    df_gem_filtered = df_gem[~df_gem.index.str.contains('handmatig', na=False)]

    # Combineer gemiddelde resultaten
    df_gem_new = DataFrame({
        'Analyse methode': df_gem_filtered.index,
        'Snijpunt y-as [kPa]': df_gem_filtered['snijpunt y-as [kPa]'].values,
        'S [-]': df_gem_filtered['Schuifsterkteratio S [-]'].values,
        'm [-]': df_gem_filtered['sterkte toename exponent = m [-]'].values,
        'POP [kPa]': df_gem_filtered['POP [kPa]'].values
    })

    # Format numeric values and handle NaN properly
    for col in df_gem_new.columns:
        if col != 'Analyse methode':
            df_gem_new[col] = df_gem_new[col].apply(
                lambda x: f"{x:.3f}" if isinstance(x, (float, int)) and x is not None and not np.isnan(x) else ""
            )

    # Maak tabel data zonder index
    header = df_gem_new.columns.tolist()
    data = [row.tolist() for _, row in df_gem_new.iterrows()]
    t_data = [header] + data

    t = LongTable(t_data, repeatRows=1, hAlign='LEFT')
    t.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return t


def _create_kar_results_table(self: "SHANSEP") -> Table:
    """
    Maakt een tabel met de SHANSEP karakteristieke resultaten (alleen voorgestelde parameters).

    Returns
    -------
    Table
        ReportLab tabel object met de SHANSEP karakteristieke resultaten
    """
    df_gem, df_kar = self.get_result_values_shansep()

    # Filter out handmatige rows - alleen voorgestelde parameters behouden
    df_kar_filtered = df_kar[~df_kar.index.str.contains('handmatig', na=False)]

    df_kar_new = DataFrame({
        'Analyse methode': df_kar_filtered.index,
        'Snijpunt y-as [kPa]': df_kar_filtered['snijpunt y-as [kPa]'].values,
        'S [-]': df_kar_filtered['Schuifsterkteratio S [-]'].values,
        'm [-]': df_kar_filtered['sterkte toename exponent = m [-]'].values,
        'POP [kPa]': df_kar_filtered['POP [kPa]'].values,
    })

    # Format numeric values and handle NaN properly
    for col in df_kar_new.columns:
        if col != 'Analyse methode':
            df_kar_new[col] = df_kar_new[col].apply(
                lambda x: f"{x:.3f}" if isinstance(x, (float, int)) and x is not None and not np.isnan(x) else ""
            )

    # Maak tabel data zonder index
    header = df_kar_new.columns.tolist()
    data = [row.tolist() for _, row in df_kar_new.iterrows()]
    t_data = [header] + data

    t = LongTable(t_data, repeatRows=1, hAlign='LEFT')
    t.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return t


def _create_handmatig_results_table(self: "SHANSEP") -> Table:
    """
    Maakt een tabel met de SHANSEP handmatige parameters in horizontale indeling.

    Returns
    -------
    Table
        ReportLab tabel object met de SHANSEP handmatige parameters
    """
    # Verzamel gemiddelde handmatige waarden
    gem_data = {
        'Snijpunt y-as [kPa]': round(self.snijpunt_gem_handmatig, 3) if self.snijpunt_gem_handmatig is not None else "",
        'S [-]': round(self.s_gem_handmatig, 3) if self.s_gem_handmatig is not None else "",
        'm [-]': round(self.m_gem_handmatig, 3) if self.m_gem_handmatig is not None else "",
        'POP [kPa]': round(self.pop_gem_handmatig, 3) if self.pop_gem_handmatig is not None else ""
    }

    # Verzamel karakteristieke handmatige waarden
    kar_data = {
        'Snijpunt y-as [kPa]': round(self.snijpunt_kar_handmatig, 3) if self.snijpunt_kar_handmatig is not None else "",
        'S [-]': round(self.s_kar_handmatig, 3) if self.s_kar_handmatig is not None else "",
        'm [-]': round(self.m_kar_handmatig, 3) if self.m_kar_handmatig is not None else "",
        'POP [kPa]': round(self.pop_kar_handmatig, 3) if self.pop_kar_handmatig is not None else ""
    }

    # Maak DataFrame in hetzelfde formaat als gem/kar tabellen
    df_handmatig = DataFrame({
        'Analyse methode': ['Gemiddeld (handmatig)', 'Karakteristiek (handmatig)'],
        'Snijpunt y-as [kPa]': [gem_data['Snijpunt y-as [kPa]'], kar_data['Snijpunt y-as [kPa]']],
        'S [-]': [gem_data['S [-]'], kar_data['S [-]']],
        'm [-]': [gem_data['m [-]'], kar_data['m [-]']],
        'POP [kPa]': [gem_data['POP [kPa]'], kar_data['POP [kPa]']]
    })

    # Maak tabel data
    header = df_handmatig.columns.tolist()
    data = [row.tolist() for _, row in df_handmatig.iterrows()]
    t_data = [header] + data

    t = LongTable(t_data, repeatRows=1, hAlign='LEFT')
    t.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return t


def save_to_pdf(self: "SHANSEP", path: str) -> str:
    """
    Slaat de SHANSEP analyseresultaten op in een PDF-document.

    De PDF bevat:
    - Titel met analysedetails
    - Beide overzichtsfiguren van de analyse (sv-su en ln(OCR)-ln(su/svc))
    - Su tabel
    - Tabel met gemiddelde en karakteristieke resultaten
    - Tabel met invoerselectie informatie

    Parameters
    ----------
    self: SHANSEP
        Instantie van de SHANSEP klasse
    path: str
        Map locatie waar het PDF-bestand moet worden opgeslagen

    Returns
    -------
    str
        Het absolute bestandspad van het aangemaakte PDF-bestand
    """
    # Maak titel en bestandsnaam
    title = f"SHANSEP analyse met {self.effective_stress} op {self.investigation_groups[0]}"
    file_name = (f"shansep_pdf_export_{self.investigation_groups[0]}_{self.analysis_type}_"
                 f"{str(self.effective_stress).replace('%', 'procent_').replace(' ', '')}.pdf")
    file_path = f"{path}/{file_name}"

    # Ensure analysis is run
    self._run_shansep()

    # Maak het PDF-document
    doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    # Voeg alleen styles toe die nog niet bestaan
    if 'Left' not in styles:
        styles.add(ParagraphStyle(name='Left', parent=styles['Normal'], alignment=TA_LEFT))
    if 'TitleLeft' not in styles:
        styles.add(ParagraphStyle(name='TitleLeft', parent=styles['Title'], alignment=TA_LEFT))
    if 'Heading3' not in styles:
        styles.add(ParagraphStyle(name='Heading3', parent=styles['Heading2'], alignment=TA_LEFT))

    # Voeg titel toe
    story = [Paragraph(title, styles['TitleLeft']), Spacer(width=1, height=12)]

    # Probeer figuren te genereren en toe te voegen
    # # Controleer of handmatige parameters zijn ingesteld
    # has_manual_params = hasattr(self, 'parameters_handmatig') and self.parameters_handmatig

    # Eerste figuur: sv-su plot
    try:
        fig_path1 = f"{path}/temp_plot1.png"
        self.show_title = False

        # Genereer figuur
        self.figure_sv_su = go.Figure()
        self.set_figure_sv_su()

        if hasattr(self, 'figure_sv_su') and self.figure_sv_su is not None:
            fig_width = 1280
            fig_height = 720
            self.figure_sv_su.write_image(fig_path1, width=fig_width, height=fig_height, scale=4, format="png")

            # Bereken figuurgrootte voor pagina 1 (met titel en heading)
            # Ruimte voor titel (ongeveer 50 pt) + heading (ongeveer 25 pt) + spacers (ongeveer 25 pt)
            title_and_heading_height = 100
            available_height = doc.height - title_and_heading_height

            # Laad PNG en bepaal pixelafmetingen
            with PILImage.open(fig_path1) as im:
                img_width_px, img_height_px = im.size

            # Bereken optimale grootte die op de pagina past
            max_width_pt = doc.width * 0.95
            aspect = img_height_px / img_width_px

            # Probeer eerst op basis van breedte
            img_width_pt = min(max_width_pt, doc.width)
            img_height_pt = img_width_pt * aspect

            # Controleer of het past in de beschikbare hoogte, anders schaal naar hoogte
            if img_height_pt > available_height:
                img_height_pt = available_height * 0.95
                img_width_pt = img_height_pt / aspect

            # Maak ReportLab Image aan
            img1 = RLImage(fig_path1)
            img1.drawWidth = img_width_pt
            img1.drawHeight = img_height_pt
            img1.hAlign = 'LEFT'

            story.append(
                Paragraph("SU-tabel (σ'<sub>v</sub> - s<sub>u</sub> relatie)", styles['Heading2']))

            story.append(Spacer(width=1, height=12))
            story.append(img1)

            # Nieuwe pagina voor tweede figuur
            story.append(PageBreak())

    except Exception as e:
        print(f"Waarschuwing: Kon eerste figuur (Sv-Su) niet toevoegen aan PDF: {e}")
        story.append(Paragraph("Eerste figuur: (kon niet worden gegenereerd)", styles['Heading2']))
        story.append(Spacer(width=1, height=12))
        story.append(PageBreak())

    # Tweede figuur: ln(OCR) - ln(su/svc) plot
    try:
        fig_path2 = f"{path}/temp_plot2.png"
        self.show_title = False

        # Genereer figuur
        self.figure_ln_ocr_ln_s = go.Figure()
        self.set_figure_ln_ocr_ln_s()

        if hasattr(self, 'figure_ln_ocr_ln_s') and self.figure_ln_ocr_ln_s is not None:
            fig_width = 1280
            fig_height = 720
            self.figure_ln_ocr_ln_s.write_image(fig_path2, width=fig_width, height=fig_height, scale=4, format="png")

            # Bereken figuurgrootte voor pagina 2 (alleen met heading)
            # Ruimte voor heading (ongeveer 25 pt) + spacer (ongeveer 12 pt)
            heading_height = 37
            available_height = doc.height - heading_height

            # Laad PNG en bepaal pixelafmetingen
            with PILImage.open(fig_path2) as im:
                img_width_px, img_height_px = im.size

            # Bereken optimale grootte die op de pagina past
            max_width_pt = doc.width * 0.95
            aspect = img_height_px / img_width_px

            # Probeer eerst op basis van breedte
            img_width_pt = min(max_width_pt, doc.width)
            img_height_pt = img_width_pt * aspect

            # Controleer of het past in de beschikbare hoogte, anders schaal naar hoogte
            if img_height_pt > available_height:
                img_height_pt = available_height * 0.95
                img_width_pt = img_height_pt / aspect

            # Maak ReportLab Image aan
            img2 = RLImage(fig_path2)
            img2.drawWidth = img_width_pt
            img2.drawHeight = img_height_pt
            img2.hAlign = 'LEFT'

            story.append(Paragraph("LN(OCR) - LN(s<sub>u</sub>/σ'<sub>v</sub>) relatie", styles['Heading2']))

            story.append(Spacer(width=1, height=12))
            story.append(img2)

            # Nieuwe pagina voor derde figuur
            story.append(PageBreak())

    except Exception as e:
        print(f"Waarschuwing: Kon tweede figuur (ln(OCR) - ln(s<sub>u</sub>/σ'<sub>v</sub>)) "
              f"niet toevoegen aan PDF: {e}")
        story.append(Paragraph("Tweede figuur: (kon niet worden gegenereerd)", styles['Heading2']))
        story.append(Spacer(width=1, height=12))
        story.append(PageBreak())

    # Derde figuur: sv-su NC plot
    try:
        fig_path3 = f"{path}/temp_plot3.png"
        self.show_title = False

        self.figure_sv_su_nc = go.Figure()
        self.set_figure_sv_su_nc()

        if hasattr(self, 'figure_sv_su_nc') and self.figure_sv_su is not None:
            fig_width = 1280
            fig_height = 720
            self.figure_sv_su_nc.write_image(fig_path3, width=fig_width, height=fig_height, scale=4, format="png")

            # Bereken figuurgrootte voor pagina 3 (alleen met heading)
            # Ruimte voor heading (ongeveer 25 pt) + spacer (ongeveer 12 pt)
            heading_height = 37
            available_height = doc.height - heading_height

            # Laad PNG en bepaal pixelafmetingen
            with PILImage.open(fig_path3) as im:
                img_width_px, img_height_px = im.size

            # Bereken optimale grootte die op de pagina past
            max_width_pt = doc.width * 0.95
            aspect = img_height_px / img_width_px

            # Probeer eerst op basis van breedte
            img_width_pt = min(max_width_pt, doc.width)
            img_height_pt = img_width_pt * aspect

            # Controleer of het past in de beschikbare hoogte, anders schaal naar hoogte
            if img_height_pt > available_height:
                img_height_pt = available_height * 0.95
                img_width_pt = img_height_pt / aspect

            # Maak ReportLab Image aan
            img3 = RLImage(fig_path3)
            img3.drawWidth = img_width_pt
            img3.drawHeight = img_height_pt
            img3.hAlign = 'LEFT'

            story.append(
                Paragraph("SU-tabel (σ'<sub>v</sub> - s<sub>u</sub> relatie) op alleen NC proeven",
                          styles['Heading2']))

            story.append(Spacer(width=1, height=12))
            story.append(img3)

            # Nieuwe pagina voor de tabellen en resultaten
            story.append(PageBreak())

    except Exception as e:
        print(f"Waarschuwing: Kon derde figuur (Sv-Su NC) niet toevoegen aan PDF: {e}")
        story.append(Paragraph("Derde figuur: (kon niet worden gegenereerd)", styles['Heading2']))
        story.append(Spacer(width=1, height=12))
        story.append(PageBreak())

    # Reset title setting
    self.show_title = True

    # Voeg parameters toe
    story.append(Paragraph("Invoer parameters en eigenschappen", styles['Heading2']))
    story.append(_create_parameters_table(self))
    story.append(Spacer(1, 12))

    # Voeg resultaten toe
    story.append(Paragraph("SHANSEP Resultaten eerste benadering gemiddeld", styles['Heading2']))
    story.append(_create_gem_results_table(self))
    story.append(Spacer(1, 12))

    # Voeg resultaten toe
    story.append(Paragraph("SHANSEP Resultaten eerste benadering karakteristiek", styles['Heading2']))
    story.append(_create_kar_results_table(self))
    story.append(Spacer(1, 12))

    story.append(PageBreak())

    # Voeg handmatige resultaten toe (alleen als er handmatige parameters zijn)
    if any([
        self.snijpunt_gem_handmatig is not None,
        self.s_gem_handmatig is not None,
        self.m_gem_handmatig is not None,
        self.pop_gem_handmatig is not None,
        self.snijpunt_kar_handmatig is not None,
        self.s_kar_handmatig is not None,
        self.m_kar_handmatig is not None,
        self.pop_kar_handmatig is not None
    ]):
        story.append(Paragraph("SHANSEP Handmatige parameters", styles['Heading2']))
        story.append(_create_handmatig_results_table(self))
        story.append(Spacer(1, 12))

    # Voeg Su tabel toe indien beschikbaar
    if hasattr(self, 'sutabel') and self.sutabel is not None:
        story.append(Paragraph("Su Tabel", styles['Heading2']))

        # Handle NaN values in the Su table
        sutabel_clean = self.sutabel.copy()

        # Replace NaN values with empty strings for all columns
        for col in sutabel_clean.columns:
            sutabel_clean[col] = sutabel_clean[col].apply(
                lambda x: "" if (isinstance(x, float) and np.isnan(x)) or x is None else x
            )
            sutabel_clean[col] = sutabel_clean[col].map(lambda x: f"{x:.3f}" if isinstance(x, (float, int)) else x)

        # Maak tabel data zonder index
        header = sutabel_clean.columns.tolist()
        data = [row.tolist() for _, row in sutabel_clean.iterrows()]
        sutabel_data = [header] + data

        sutabel_table = LongTable(sutabel_data, repeatRows=1, hAlign='LEFT')
        sutabel_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(sutabel_table)
        story.append(Spacer(1, 12))

    # Voeg invoerselectie toe
    story.append(Paragraph("Invoerselectie Informatie", styles['Heading2']))
    story.append(_create_input_table(self))

    # Maak PDF
    doc.build(story)
    print(f"SHANSEP PDF export voltooid: {file_path}")

    # Ruim tijdelijke plot bestanden op
    import os
    temp_files = [f"{path}/temp_plot1.png", f"{path}/temp_plot2.png", f"{path}/temp_plot3.png"]
    for temp_file in temp_files:
        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except Exception as e:
            print(f"Waarschuwing: Kon tijdelijk plot bestand niet verwijderen {temp_file}: {e}")

    return file_path
