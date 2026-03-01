from openpyxl import load_workbook
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
from openpyxl.utils import get_column_letter


def format_excel_sheet(
    file_path: str,
    sheet_name: str,
    num_columns: int,
    num_rows: int,
    table_name: str = None,
    index: bool = True
):
    """
    Format an Excel worksheet as a table with filters and custom column widths.

    Parameters
    ----------
    file_path : str
        Full path to the Excel file
    sheet_name : str
        Name of the worksheet to format
    num_columns : int
        Number of data columns (excluding the index)
    num_rows : int
        Number of rows in the table (excluding the header)
    table_name : str, optional
        Name for the Excel table. If None, will use sheet_name + "Table".
    index : bool, default True
        Whether the index column is present
    """

    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]

    # Adjust for index column if present
    total_columns = num_columns + (1 if index else 0)

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Define table range
    table_range = f"A1:{get_column_letter(total_columns)}{num_rows + 1}"

    # Create table with filters
    if table_name is None:
        table_name = f"{sheet_name}Table"

    # Remove spaces and special characters from table name
    table_name = "".join(c for c in table_name if c.isalnum())

    table = XLTable(displayName=table_name, ref=table_range)

    # Add a default style
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style

    # Remove existing table if it exists
    for existing_table in list(worksheet.tables.values()):
        if existing_table.name == table_name:
            del worksheet.tables[existing_table.name]
            break

    # Add the table to the worksheet
    worksheet.add_table(table)
    workbook.save(file_path)
