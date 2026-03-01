from pandas import DataFrame
import pandas as pd
from openpyxl import load_workbook
import importlib.resources
from typing import Optional, Literal
from pathlib import Path
import os.path

from pv_tool.imports.create_dbase import add_missing_columns, select_columns, alg_columns, add_ana_columns, add_pv_naam
from pv_tool.imports.import_options import import_dbase, import_pv_tool, import_stowa
from pv_tool.imports.validation import Validation
from pv_tool.imports.globals import PV_TOOL_DBASE_COLUMNS, ANA_COLUMNS


class Dbase:
    """Deze class bevat alle functies die te maken hebben met het bouwen de Dbase-dataframe"""

    def __init__(self):
        self.stowa_df: Optional[DataFrame] = None
        self.pv_tool: Optional[DataFrame] = None
        self.dbase_df: Optional[DataFrame] = None
        self.validation = Validation(dbase=self)

    def _create_dbase(self, source: Literal['Stowa', 'PV-tool', 'Dbase']):
        """Maakt de dbase-dataframe"""
        if source == 'Stowa':
            add_missing_columns(self)
            alg_columns(self)
            add_ana_columns(self)
            add_pv_naam(self)
        elif source == 'PV-tool':
            select_columns(self)
            alg_columns(self)
            add_ana_columns(self)
            add_pv_naam(self)
        elif source == 'Dbase':
            add_ana_columns(self)
            add_pv_naam(self)

    def import_dbase_short(self, source: Literal['Stowa', 'PV-tool', 'Dbase'], source_dir: Path):
        """Importeert data uit de Stowa-database, de oude pv-tool of de Dbase (template) en voegt kolommen toe"""
        if source == 'Dbase':
            import_dbase(self, dbase_dir=source_dir)
            return self.dbase_df
        else:
            return f"Short import only available for 'Dbase' source, not for '{source}'"

    def import_data(self, source: Literal['Stowa', 'PV-tool', 'Dbase'], source_dir: Path):
        """Importeert data uit de Stowa-database, de oude pv-tool of de Dbase (template) en voegt kolommen toe"""
        if source == 'Stowa':
            import_stowa(self, stowa_dir=source_dir)
            self.dbase_df = self.stowa_df
        elif source == 'PV-tool':
            import_pv_tool(self, pv_dir=source_dir)
            self.dbase_df = self.pv_tool
        elif source == 'Dbase':
            import_dbase(self, dbase_dir=source_dir)
        self._create_dbase(source=source)
        return self.dbase_df

    def validate_data(self, export_path: Path):
        self.validation.validation_export(export_path=export_path)
        self.validation.print_critical_errors()
        return self.dbase_df

    def create_dbase_for_export(self):
        """
        Creates the Dbase-DataFrame for the export to Excel-template, maintaining the correct column order
        and preserving specified columns from the current DataFrame.
        """
        # Reorder columns for template
        base_columns = [col for col in PV_TOOL_DBASE_COLUMNS if col in self.dbase_df.columns]
        base_columns = [col for col in base_columns if col not in ANA_COLUMNS]
        ana_columns = [col for col in ANA_COLUMNS if col in self.dbase_df.columns]
        used_columns = set(base_columns + ana_columns)
        other_columns = [col for col in self.dbase_df.columns if col not in used_columns]
        final_columns = base_columns + ana_columns + other_columns

        # Save in init with reordered columns
        return self.dbase_df[final_columns].copy()

    def export_dbase_to_template(self, export_dir, export_name="Template_PVtool5_0.xlsx"):
        """Exporteert het dbase-dataframe naar de excel-template"""

        sheet_name = 'Dbase5_0'
        start_row = 7  # Excel: rij 8
        start_col = 1  # Excel: kolom A

        with importlib.resources.path('pv_tool.templates', "Template_PVtool5_0.xlsx") as template_path:
            wb = load_workbook(template_path)

        # wb = load_workbook(template_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' bestaat niet in template!")

        ws = wb[sheet_name]

        export_df = self.create_dbase_for_export()
        index_col_name = export_df.index.name if export_df.index.name else "Index"
        export_df.insert(0, index_col_name, export_df.index)

        export_df = export_df.replace({pd.NA: ""})  # want kan geen NA omschrijven naar excel

        for i, row in enumerate(export_df.values):
            for j, value in enumerate(row):
                ws.cell(row=start_row + 1 + i, column=start_col + j, value=value)

        export_to = os.path.join(export_dir, export_name)
        wb.save(export_to)
        print(f"DataFrame naar template geëxporteerd in {export_to}")
