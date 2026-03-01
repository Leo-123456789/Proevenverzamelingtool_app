from typing import Optional, List, Literal
from pathlib import Path
from datetime import datetime
from pv_tool.utilities.utils import get_repo_root
from pandas import DataFrame, concat, read_excel, isna
from pv_tool.cphi_analysis.globals import (TEXTUAL_NAMES, ALL_TEXTUAL_NAMES,
                                           NEW_COLUMN_NAMES, TEXTUAL_NAMES_DSS, ALL_TEXTUAL_NAMES_DSS)
from pv_tool.cphi_analysis.save_and_export import save_total_to_excel, save_to_pdf

from pv_tool.imports.import_data import Dbase
import plotly.graph_objects as go
from pv_tool.cphi_analysis.expand_analysis_df import (calculate_tan_a, calculate_ln_tan_a, calculate_s_tt,
                                                      calculate_s_ty, calculate_kappa_2, calculate_s,
                                                      calculate_5pr_ondergrens, calculate_5pr_bovengrens,
                                                      calculate_s_tt_ondergrens, calculate_s_ty_ondergrens,
                                                      calculate_kappa_2_ondergrens, calculate_correctie_t,
                                                      kappa_2_2pr_cor,
                                                      calculate_5pr_ondergrens_correctie_c,
                                                      calculate_5pr_bovengrens_correctie_c,
                                                      calculate_s_ty_ondergrens_correctie_c,
                                                      calculate_kappa_2_ondergrens_correctie_c)
from pv_tool.cphi_analysis.visualization import (add_proefresultaten, add_extra_proefresultaten, add_5pr_bovengrens,
                                                 add_5pr_ondergrens, add_fysische_realiseerbare_ondergrens,
                                                 add_gemiddelde,
                                                 set_layout, add_gemiddelde_sh, add_raaklijn_kar_boven,
                                                 add_raaklijn_kar_onder)
from pv_tool.cphi_analysis.calc_parameters import (calc_watergehalte_gem, calc_watergehalte_sd, calc_vgwnat_gem,
                                                   calc_vgwnat_sd, calc_a2_phi_gem, calc_a2_kar, calc_phi_d,
                                                   helling_gecorrigeerd, calc_a1_c_gem, calc_tan_phi_gem, calc_phi_kar,
                                                   calc_a1_kar, calc_phi_gem, calc_c_gem, calc_tan_phi_kar,
                                                   calc_c_kar, calc_tan_phi_d, calc_c_d, calc_st_dev_phi, calc_st_dev_c,
                                                   calc_a2_phi_gem_sh, calc_a2_phi_kar_boven_sh,
                                                   calc_a2_phi_kar_onder_sh,
                                                   calc_tan_phi_kar_sh)
from openpyxl import load_workbook


class CPhiAnalyse:
    """
    Klasse voor het uitvoeren van c-phi analyses op grondmonsters.

    Ondersteunt zowel triaxiaal (TXT) als direct simple shear (DSS) testen,
    en kan zowel reguliere c-phi als schematiseringshandleiding (SH) analyses uitvoeren.
    """

    def __init__(self, dbase: Dbase,
                 analysis_type: Literal['TXT_CPhi', 'TXT_SH', 'DSS_CPhi', 'DSS_SH'],
                 investigation_groups: List,
                 effective_stress: Literal[
                     '2% rek', '5% rek', '10% rek', '15% rek', '20% rek', 'pieksterkte', 'eindsterkte']):
        """
        Initialiseert een nieuwe c-phi analyse.

        Parameters
        ----------
        dbase : Dbase
            Database object met proefresultaten
        analysis_type : str
            Type analyse ('TXT_CPhi', 'TXT_SH', 'DSS_CPhi', 'DSS_SH')
        investigation_groups : List
            Lijst met te analyseren proefgroepen
        effective_stress : str
            Rekpercentage of sterkte criterium voor de analyse
        """

        # Validate effective_stress based on analysis_type
        if analysis_type in ['TXT_CPhi', 'TXT_SH'] and effective_stress in ['10% rek', '20% rek']:
            raise ValueError(
                f"De waardes '10% rek' and '20% rek' kunnen alleen gebruikt worden bij de DSS analyse. "
                f"De gekozen analyse is: '{analysis_type}', en de sterkte is: '{effective_stress}'"
            )

        # Data
        self.dbase_df = dbase.dbase_df
        self.analysis_type = analysis_type
        self.investigation_groups = investigation_groups
        self.effective_stress = effective_stress

        # Settings
        self.alpha: Optional[float] = 0.75
        self.material_cohesie: Optional[float] = 1.0
        self.material_tan_phi: Optional[float] = 1.0

        # Parameters
        self.eerste_benadering_a2_gem: Optional[float] = None  # phi_gem
        self.eerste_benadering_a2_kar: Optional[float] = None  # phi_kar
        self.eerste_benadering_a1_gem: Optional[float] = None  # cohesie_gem
        self.eerste_benadering_a1_kar: Optional[float] = None  # cohesie_kar

        self.helling_gecorrigeerd: Optional[float] = None
        self.cohesie_gem_handmatig: Optional[float] = None
        self.phi_kar_handmatig: Optional[float] = None
        self.cohesie_kar_handmatig: Optional[float] = None

        self.calc_watergehalte_gem: Optional[float] = None
        self.calc_watergehalte_sd: Optional[float] = None
        self.calc_vgwnat_gem: Optional[float] = None
        self.calc_vgwnat_sd: Optional[float] = None

        # Placeholder
        self.cphi_analyses_data_df: Optional[DataFrame] = None
        self.total_cphi_analyses_data_df: Optional[DataFrame] = None

        # Results
        self.tan_phi_gem: Optional[float] = None
        self.phi_gem: Optional[float] = None
        self.c_gem: Optional[float] = None
        self.tan_phi_kar: Optional[float] = None
        self.phi_kar: Optional[float] = None
        self.c_kar: Optional[float] = None
        self.tan_phi_d: Optional[float] = None
        self.phi_d: Optional[float] = None
        self.c_d: Optional[float] = None
        self.st_dev_phi: Optional[float] = None
        self.st_dev_c: Optional[float] = None

        self.gem_a1: Optional[float] = None
        self.gem_a2: Optional[float] = None

        self.kar_a1: Optional[float] = None
        self.kar_a2: Optional[float] = None

        self.a2_phi_kar_onder: Optional[float] = None
        self.a2_phi_kar_boven: Optional[float] = None

        # Figure
        self.figure = go.Figure()
        self.show_title: Optional[bool] = True

    # ========= Instelling en Data Ophalen Methodes ==========

    def get_cphi_data(self):
        """
        Filtert de database op basis van analysetype en proefgroepen.

        Selecteert de juiste data voor de analyse op basis van het type test (TXT/DSS),
        de proefgroepen en het gewenste rekpercentage. Berekent tevens gemiddelde
        eigenschappen zoals watergehalte.
        """
        if self.analysis_type in ['TXT_CPhi', 'TXT_SH']:
            self.cphi_analyses_data_df = self.dbase_df[self.dbase_df['ALG__TRIAXIAAL']]
            self.cphi_analyses_data_df = self.cphi_analyses_data_df[self.cphi_analyses_data_df['PV_NAAM'].isin(
                self.investigation_groups)]
            self.calc_watergehalte_gem = calc_watergehalte_gem(self)
            self.calc_watergehalte_sd = calc_watergehalte_sd(self)
            self.calc_vgwnat_gem = calc_vgwnat_gem(self)
            self.calc_vgwnat_sd = calc_vgwnat_sd(self)
            self.total_cphi_analyses_data_df = self.cphi_analyses_data_df
            self.cphi_analyses_data_df = self.cphi_analyses_data_df[TEXTUAL_NAMES.get(self.effective_stress, [])]

            # Valideer of er data overblijft na filtering
            if self.cphi_analyses_data_df.empty:
                raise ValueError(f"Geen data gevonden na filtering op investigation_groups {self.investigation_groups} "
                                 f"en effective_stress '{self.effective_stress}' "
                                 f"voor analyse type '{self.analysis_type}'")

            print(f"Data na filtering: {len(self.cphi_analyses_data_df)} rijen gevonden")

        elif self.analysis_type in ['DSS_CPhi', 'DSS_SH']:
            self.cphi_analyses_data_df = self.dbase_df[self.dbase_df['ALG__DSS']]
            self.cphi_analyses_data_df = self.cphi_analyses_data_df[self.cphi_analyses_data_df['PV_NAAM'].isin(
                self.investigation_groups)]
            self.calc_watergehalte_gem = calc_watergehalte_gem(self)
            self.calc_watergehalte_sd = calc_watergehalte_sd(self)
            self.calc_vgwnat_gem = calc_vgwnat_gem(self)
            self.calc_vgwnat_sd = calc_vgwnat_sd(self)
            self.total_cphi_analyses_data_df = self.cphi_analyses_data_df
            self.cphi_analyses_data_df = self.cphi_analyses_data_df[TEXTUAL_NAMES_DSS.get(self.effective_stress, [])]

            # Valideer of er data overblijft na filtering
            if self.cphi_analyses_data_df.empty:
                raise ValueError(f"Geen data gevonden na filtering op investigation_groups {self.investigation_groups} "
                                 f"en effective_stress '{self.effective_stress}' "
                                 f"voor analyse type '{self.analysis_type}'")

            print(f"Data na filtering: {len(self.cphi_analyses_data_df)} rijen gevonden")

        self.cphi_analyses_data_df.columns = NEW_COLUMN_NAMES

    def apply_settings(self, alpha: Optional[float] = None,
                       material_factor_cohesion: Optional[float] = None,
                       material_factor_tan_phi: Optional[float] = None):
        """
        Past analyse-instellingen aan.

        Parameters
        ----------
        alpha : float, optioneel
            Alpha-waarde voor de analyse
        material_factor_cohesion : float, optioneel
            Materiaalfactor voor cohesie
        material_factor_tan_phi : float, optioneel
            Materiaalfactor voor tan(phi)
        """
        self.alpha = alpha if alpha is not None else self.alpha
        self.material_cohesie = material_factor_cohesion if material_factor_cohesion is not None \
            else self.material_cohesie
        self.material_tan_phi = material_factor_tan_phi if material_factor_tan_phi is not None \
            else self.material_tan_phi

    def apply_parameters(self, cohesie_gem: Optional[float] = None,
                         phi_kar: Optional[float] = None,
                         cohesie_kar: Optional[float] = None):
        """
        Past handmatige parameters aan en herberekent de analyse.

        Parameters
        ----------
        cohesie_gem : float, optioneel
            Gemiddelde cohesie
        phi_kar : float, optioneel
            Karakteristieke phi-waarde
        cohesie_kar : float, optioneel
            Karakteristieke cohesie
        """
        if cohesie_gem is not None:
            self.cohesie_gem_handmatig = cohesie_gem
        if phi_kar is not None:
            self.phi_kar_handmatig = phi_kar
        if cohesie_kar is not None:
            self.cohesie_kar_handmatig = cohesie_kar

        if self.analysis_type in ['TXT_SH', 'DSS_SH']:
            self._run_sh()
        else:
            self._run()

    def plot_spanningspaden(self):
        """
        Initieert de spanningspaden voor alle beschikbare effective stress waarden
        binnen de geselecteerde investigation groups.

        Dit helpt bij het visualiseren van de spanningsveranderingen
        tijdens de proeven. Voor elk monster (uit de index van de dataframes)
        wordt een apart spanningspad gemaakt met alle beschikbare spanningsstappen.

        Wordt aangeroepen binnen set_figure(). Het is aan de gebruiker om de spanningspaden wel of niet toe te voegen
        aan de figuur.
        """
        if self.analysis_type in ['TXT_CPhi', 'TXT_SH']:
            relevant_df = self.dbase_df[self.dbase_df['ALG__TRIAXIAAL']]
            relevant_df = relevant_df[relevant_df['PV_NAAM'].isin(self.investigation_groups)]
            effective_stress_options = ['consolidatie', '2% rek', '5% rek', '15% rek', 'pieksterkte', 'eindsterkte']
        elif self.analysis_type in ['DSS_CPhi', 'DSS_SH']:
            relevant_df = self.dbase_df[self.dbase_df['ALG__DSS']]
            relevant_df = relevant_df[relevant_df['PV_NAAM'].isin(self.investigation_groups)]
            effective_stress_options = ['consolidatie', '2% rek', '5% rek', '10% rek', '15% rek', '20% rek',
                                        'pieksterkte', 'eindsterkte']
            relevant_df['DSS_T_CONSOLIDATIE'] = [0] * len(relevant_df)
        else:
            raise ValueError("Ongeldig analysetype. Gebruik 'TXT_CPhi', 'TXT_SH', 'DSS_CPhi' of 'DSS_SH'.")

        # Eerst verzamelen we alle data per spanningsstap
        all_data = {}
        for stress in effective_stress_options:
            columns = ALL_TEXTUAL_NAMES.get(stress, []) \
                if self.analysis_type in ['TXT_CPhi', 'TXT_SH'] else ALL_TEXTUAL_NAMES_DSS.get(stress, [])
            if len(columns) > 0:
                data = relevant_df[columns].copy()
                if not data.empty and not data.isna().all().all():
                    data.columns = NEW_COLUMN_NAMES
                    all_data[stress] = data

        # Nu herstructureren we de data per monster (uit de index)
        sample_stress_paths = {}

        # Verzamel alle unieke monster namen uit de indices van alle dataframes
        all_samples = set()
        for df in all_data.values():
            all_samples.update(df.index)

        # Voor elk monster, verzamel alle spanningsstappen
        for sample_name in all_samples:
            stress_data = {'S\'': [], 'T': [], 'stress_state': []}
            for stress, df in all_data.items():
                if sample_name in df.index:
                    sample_data = df.loc[sample_name]
                    if not (isna(sample_data['S\'']) or isna(sample_data['T'])):
                        stress_data['S\''].append(sample_data['S\''])
                        stress_data['T'].append(sample_data['T'])
                        stress_data['stress_state'].append(stress)

            if stress_data['S\'']:  # Als er data is voor dit monster
                stress_df = DataFrame(stress_data)

                if self.analysis_type in ['TXT_CPhi', 'TXT_SH']:
                    rek_bij_t_piek = self.total_cphi_analyses_data_df.loc[sample_name, 'TXT_SS_REK_BIJ_T_PIEK']
                    rek_bij_t_eind = self.total_cphi_analyses_data_df.loc[sample_name, 'TXT_SS_REK_BIJ_T_EIND']
                    if not isna(rek_bij_t_piek) and 'pieksterkte' in stress_df['stress_state'].values:
                        # Verplaats de rij met 'pieksterkte' naar de juiste positie
                        piek_row = stress_df[stress_df['stress_state'] == 'pieksterkte']
                        stress_df = stress_df[stress_df['stress_state'] != 'pieksterkte']
                        if rek_bij_t_piek < 2:
                            insert_index = 1
                        elif rek_bij_t_piek < 5:
                            insert_index = 2
                        elif rek_bij_t_piek < 15:
                            insert_index = 3
                        elif not isna(rek_bij_t_eind) and rek_bij_t_eind > rek_bij_t_piek:
                            insert_index = 4
                        else:
                            insert_index = len(stress_df)
                        stress_df = concat(
                            [stress_df.iloc[:insert_index], piek_row, stress_df.iloc[insert_index:]]).reset_index(
                            drop=True)
                elif self.analysis_type in ['DSS_CPhi', 'DSS_SH']:
                    rek_bij_t_max = self.total_cphi_analyses_data_df.loc[sample_name, 'DSS_REK_BIJ_T_MAX']
                    rek_bij_t_eind = self.total_cphi_analyses_data_df.loc[sample_name, 'DSS_REK_BIJ_T_EIND']
                    if not isna(rek_bij_t_max) and 'pieksterkte' in stress_df['stress_state'].values:
                        # Verplaats de rij met 'pieksterkte' naar de juiste positie
                        piek_row = stress_df[stress_df['stress_state'] == 'pieksterkte']
                        stress_df = stress_df[stress_df['stress_state'] != 'pieksterkte']
                        if rek_bij_t_max < 2:
                            insert_index = 1
                        elif rek_bij_t_max < 5:
                            insert_index = 2
                        elif rek_bij_t_max < 10:
                            insert_index = 3
                        elif rek_bij_t_max < 15:
                            insert_index = 4
                        elif rek_bij_t_max < 20:
                            insert_index = 5
                        elif not isna(rek_bij_t_eind) and rek_bij_t_eind > rek_bij_t_max:
                            insert_index = 6
                        else:
                            insert_index = len(stress_df)
                        stress_df = concat(
                            [stress_df.iloc[:insert_index], piek_row, stress_df.iloc[insert_index:]]).reset_index(
                            drop=True)

                sample_stress_paths[sample_name] = stress_df

        # Plot de spanningspaden
        if sample_stress_paths:
            from pv_tool.cphi_analysis.visualization import add_stress_paths
            add_stress_paths(self, sample_stress_paths)
        else:
            raise ValueError("Geen geldige data gevonden voor de spanningspaden.")

    def get_previous_results(self, path: str, file_name: str):
        """
        Zoekt naar eerdere analyseresultaten in een Excel-bestand.

        Parameters
        ----------
        path: str
            Pad naar de map waar het Excel-bestand staat
        file_name: str
            Naam van het Excel-bestand

        Returns
        -------
        DataFrame of None
            DataFrame met eerdere resultaten als deze gevonden zijn, anders None
        """

        file_path = f"{path}/{file_name}"

        try:
            with open(file_path, 'r'):
                pass
        except FileNotFoundError:
            raise FileNotFoundError("Er is geen dbase aanwezig onder de naam Template_PVtool5_0.xlsx")

        try:
            results_df = read_excel(file_path, sheet_name='Resultaten c-phi', skiprows=6)
        except ValueError:
            print("Er is geen tabblad 'Resultaten c-phi' aanwezig in het Excel-bestand.")
            return None

        filtered_df = results_df[
            (results_df['PV_RESULTAAT_ID'].str.contains(self.investigation_groups[0])) &
            (results_df['PV_RESULTAAT_ID'].str.contains(self.effective_stress)) &
            (results_df['PV_RESULTAAT_ID'].str.contains(self.analysis_type))
            ]

        if filtered_df.empty:
            print("Er zijn geen eerdere resultaten gevonden voor de opgegeven parameters.")
            return None

        latest_entry = filtered_df.sort_values(by='Timestamp', ascending=False).iloc[0]

        return latest_entry

    # ========= Analyse Methodes ==========

    def expand_analysis_df(self):
        """
        Berekent afgeleide parameters voor de c-phi analyse.

        Voegt kolommen toe aan het dataframe met berekende waarden voor
        s_tt, s_ty, kappa_2 en verschillende onder- en bovengrenzen.
        """
        calculate_s_tt(self)
        calculate_s_ty(self)
        calculate_kappa_2(self)
        calculate_s(self)
        calculate_5pr_ondergrens(self)
        calculate_5pr_bovengrens(self)
        calculate_s_tt_ondergrens(self)
        calculate_s_ty_ondergrens(self)
        calculate_kappa_2_ondergrens(self)

    def expand_analysis_df_sh(self):
        """
        Berekent afgeleide parameters voor de schematiseringshandleiding analyse.

        Voegt kolommen toe aan het dataframe met berekende waarden voor
        tan(alpha) en ln(tan(alpha)).
        """
        calculate_tan_a(self)
        calculate_ln_tan_a(self)

    def eerste_benadering(self):
        """
        Maakt een eerste schatting van de gemiddelde sterkteparameters cohesie (a1) en phi (a2).
        """
        # self.eerste_benadering_a1_gem = calc_cohesie_gem(self)
        self.eerste_benadering_a1_gem = calc_a1_c_gem(self)
        self.eerste_benadering_a2_gem = calc_a2_phi_gem(self)

    def eerste_benadering_deel2(self):
        """
        Maakt een eerste schatting van de karakteristieke sterkteparameters cohesie (a1) en phi (a2).
        """
        self.eerste_benadering_a2_kar = calc_a2_kar(self)
        self.eerste_benadering_a1_kar = calc_a1_kar(self)

    def expand_analysis_df_corrected(self):
        """
        Voegt gecorrigeerde parameters toe aan de analyse: gecorrigeerde waarden voor
        onder- en bovengrenzen en kappa_2.
        """
        calculate_correctie_t(self)
        kappa_2_2pr_cor(self)
        calculate_5pr_ondergrens_correctie_c(self)
        calculate_5pr_bovengrens_correctie_c(self)
        calculate_s_ty_ondergrens_correctie_c(self)
        calculate_kappa_2_ondergrens_correctie_c(self)

    # ========= Resultaten Methodes ==========

    def result_values(self):
        """
        Berekent de definitieve resultaten van de c-phi analyse: gemiddelde, karakteristieke en rekenwaarden voor
        cohesie en phi, inclusief standaarddeviaties.
        """
        self.helling_gecorrigeerd = helling_gecorrigeerd(self)

        self.gem_a1 = self.cohesie_gem_handmatig if self.cohesie_gem_handmatig is not None else calc_a1_c_gem(self)
        self.gem_a2 = calc_a2_phi_gem(self)

        self.kar_a1 = self.cohesie_kar_handmatig if self.cohesie_kar_handmatig is not None else calc_a1_kar(self)
        self.kar_a2 = self.phi_kar_handmatig if self.phi_kar_handmatig is not None else calc_a2_kar(self)

        self.phi_gem = calc_phi_gem(self)
        self.tan_phi_gem = calc_tan_phi_gem(self)
        self.c_gem = calc_c_gem(self)

        self.tan_phi_kar = calc_tan_phi_kar(self)
        self.phi_kar = calc_phi_kar(self)
        self.c_kar = calc_c_kar(self)

        self.phi_d = calc_phi_d(self)
        self.tan_phi_d = calc_tan_phi_d(self)
        self.c_d = calc_c_d(self)

        self.st_dev_phi = calc_st_dev_phi(self)
        self.st_dev_c = calc_st_dev_c(self)

    def result_values_sh(self):
        """
        Berekent de definitieve resultaten van de c-phi schematiseringshandleiding analyse:
        gemiddelde, karakteristieke en rekenwaarden voor phi,
        inclusief standaarddeviatie.

        """
        self.gem_a2 = calc_a2_phi_gem_sh(self)
        self.a2_phi_kar_onder = calc_a2_phi_kar_onder_sh(self)
        self.a2_phi_kar_boven = calc_a2_phi_kar_boven_sh(self)
        self.kar_a2 = self.a2_phi_kar_onder

        self.tan_phi_gem = calc_tan_phi_gem(self)
        self.tan_phi_kar = calc_tan_phi_kar_sh(self)
        self.tan_phi_d = calc_tan_phi_d(self)

        self.phi_gem = calc_phi_gem(self)
        self.phi_kar = calc_phi_kar(self)
        self.phi_d = calc_phi_d(self)

        self.st_dev_phi = calc_st_dev_phi(self)

    def _run(self):
        """
        Voert de volledige c-phi analyse uit in de juiste volgorde:
        data ophalen, parameters berekenen en resultaten bepalen.
        """
        self.get_cphi_data()
        self.expand_analysis_df()
        self.eerste_benadering()
        self.expand_analysis_df_corrected()
        self.eerste_benadering_deel2()
        self.result_values()

    def _run_sh(self):
        """
        Voert de volledige c-phi schematiseringshandleiding analyse uit in de juiste volgorde:
        data ophalen, parameters berekenen en resultaten bepalen.
        """
        self.get_cphi_data()
        self.expand_analysis_df_sh()
        self.result_values_sh()

    # ========== Visualisatie Methodes ==========

    def set_figure(self, plot_extra_dataset: Optional[List] = None, plot_spanningspaden: bool = False):
        """
        Maakt een visualisatie van de analyseresultaten.

        Parameters
        ----------
        plot_extra_dataset : List, optioneel
            Extra dataset om in de plot weer te geven

        plot_spanningspaden : bool, optioneel
            Of de spanningspaden moeten worden weergegeven
        """
        if plot_spanningspaden:
            self.plot_spanningspaden()

        add_proefresultaten(self)
        if plot_extra_dataset is not None:
            add_extra_proefresultaten(self, plot_extra_dataset)

        if self.analysis_type in ['TXT_SH', 'DSS_SH']:
            add_gemiddelde_sh(self)
            add_raaklijn_kar_onder(self)
            add_raaklijn_kar_boven(self)
        else:
            add_5pr_bovengrens(self)
            add_5pr_ondergrens(self)
            add_fysische_realiseerbare_ondergrens(self)
            add_gemiddelde(self)

        set_layout(self)

    def show_figure(self, plot_extra_dataset: Optional[List] = None, plot_spanningspaden: bool = False):
        """
        Toont de visualisatie van de analyseresultaten.

        Parameters
        ----------
        plot_extra_dataset : List, optioneel
            Extra dataset om in de plot weer te geven

        plot_spanningspaden : bool, optioneel
            Of de spanningspaden moeten worden weergegeven
        """
        if self.analysis_type in ['TXT_SH', 'DSS_SH']:
            self._run_sh()
        else:
            self._run()
        self.figure = go.Figure()
        self.set_figure(plot_extra_dataset, plot_spanningspaden=plot_spanningspaden)
        self.figure.show()

    def save_fig_html(self, path: str, export_name: Optional[str] = None):
        """
        Slaat de visualisatie van de analyseresultaten op als een HTML-bestand.

        N.B. als de figuur leeg is, roep dan eerst show_figure() aan om de figuur te genereren.

        Parameters
        ----------
        path: str
            Pad naar de map waar het bestand opgeslagen moet worden
        export_name: str, optioneel
            Naam van het HTML-bestand
        """
        if export_name is None:
            export_name = f"c-phi_analyse_{self.investigation_groups[0].replace(' ', '_')}.html"
        file_path = Path(path) / export_name
        self.figure.write_html(file_path)
        print(f"figuur opgeslagen als HTML: {file_path}")

    def get_short_results(self):
        """
        Genereert een samenvattend overzicht van de analyseresultaten.

        Returns
        -------
        DataFrame
            DataFrame met verwachtingswaarden, karakteristieke waarden,
            rekenwaarden en standaarddeviaties
        """
        if self.analysis_type in ['TXT_SH', 'DSS_SH']:
            self._run_sh()
            index = ['Verwachtingswaarde', 'Karakteristieke waarde', 'Rekenwaarde', 'Standaarddeviatie D-stability']
            columns = ['tan phi [-]', 'phi [graden]']
            analyse_output_df = DataFrame(index=index, columns=columns)
            analyse_output_df['tan phi [-]'] = [self.tan_phi_gem, self.tan_phi_kar, self.tan_phi_d, '[-]']
            analyse_output_df['phi [graden]'] = [self.phi_gem, self.phi_kar, self.phi_d, self.st_dev_phi]
        else:
            self._run()
            index = ['Verwachtingswaarde', 'Karakteristieke waarde', 'Rekenwaarde', 'Standaarddeviatie D-stability']
            columns = ['tan phi [-]', 'phi [graden]', 'cohesie [kPa]']
            analyse_output_df = DataFrame(index=index, columns=columns)
            analyse_output_df['tan phi [-]'] = [self.tan_phi_gem, self.tan_phi_kar, self.tan_phi_d, '[-]']
            analyse_output_df['phi [graden]'] = [self.phi_gem, self.phi_kar, self.phi_d, self.st_dev_phi]
            analyse_output_df['cohesie [kPa]'] = [self.c_gem, self.c_kar, self.c_d, self.st_dev_c]
        return analyse_output_df

    # ========== Export Methodes ==========

    def add_results_to_template(self, path, export_name=None):
        """
        Voegt een nieuwe resultatenrij toe aan tabblad 'Resultaten c-phi' in het Excel-template.
        Als het tabblad niet bestaat, wordt het aangemaakt en worden de kolomnamen weggeschreven.
        """
        if export_name is None:
            export_name = "Template_PVtool5_0.xlsx"
        file_path = Path(path) / export_name
        sheet_name = 'Resultaten c-phi'

        expected_columns = [
            'PV_RESULTAAT_ID', 'PV_NAAM', 'PV_REK', 'PV_TYPE_PROEF', 'PV_ANALYSE',
            'PV_A1_COH_GEM', 'PV_A2_TAN_PHI_GEM', 'PV_A1_COH_KAR', 'PV_A2_TAN_PHI_KAR',
            'PV_PARTPHI', 'PV_PARTCOH', 'PV_TYPEVERZAMELING', 'PV_COH_GEM', 'PV_PHI_GEM',
            'PV_COH_KAR', 'PV_PHI_KAR', 'PV_COH_SD_DSTAB', 'PV_PHI_SD_DSTAB',
            'PV_VGWNAT_GEM', 'PV_VGWNAT_SD', 'PV_WATERGEHALTE_GEM', 'PV_WATERGEHALTE_SD', 'Timestamp'
        ]

        # Maak de resultaat-rij aan
        new_row = {
            'PV_NAAM': self.investigation_groups[0],
            'PV_REK': self.effective_stress,
            'PV_TYPE_PROEF': self.analysis_type.split('_')[0],
            'PV_ANALYSE': self.analysis_type.split('_')[1],
            'PV_RESULTAAT_ID': f"{self.investigation_groups[0]}_{self.effective_stress}_"
                               f"{self.analysis_type.split('_')[0]}_{self.analysis_type.split('_')[1]}",
            'PV_TYPEVERZAMELING': self.alpha,
            'PV_A1_COH_GEM': round(self.gem_a1, 3) if self.gem_a1 is not None else None,
            # TODO: gaat deze dan wel goed?
            'PV_A2_TAN_PHI_GEM': round(self.gem_a2, 3) if self.gem_a2 is not None else None,
            # TODO: welke waarde? gem_a2 of tan_phi_gem
            'PV_A1_COH_KAR': round(self.kar_a1, 3) if self.kar_a1 is not None else None,
            # TODO: gaat deze dan wel goed?
            'PV_A2_TAN_PHI_KAR': round(self.kar_a2, 3) if self.kar_a2 is not None else None,
            # TODO: zelfde geld hier met kar_a2 of tan_phi_kar

            'PV_COH_GEM': (
                round(self.c_gem, 3) if self.c_gem is not None and self.c_gem >= 0
                else "[-]" if self.c_gem is None
                else f"{round(self.c_gem, 3)} (kan niet - aanpassen!)"
            ),
            'PV_PHI_GEM': round(self.phi_gem, 3) if self.phi_gem is not None else None,

            'PV_COH_KAR': (
                round(self.c_kar, 3) if self.c_kar is not None and self.c_kar >= 0
                else "[-]" if self.c_kar is None
                else f"{round(self.c_kar, 3)} (kan niet - aanpassen!)"
            ),
            'PV_PHI_KAR': round(self.phi_kar, 3) if self.phi_kar is not None else None,

            'PV_COH_SD_DSTAB': (
                round(self.st_dev_c, 3) if (
                        self.c_gem is not None and self.c_kar is not None and self.c_gem >= 0 and self.c_kar >= 0
                ) else "[-]" if self.c_gem is None or self.c_kar is None
                else "[-] (c < 0)"
            ),
            'PV_PHI_SD_DSTAB': round(self.st_dev_phi, 3) if self.st_dev_phi is not None else None,

            'PV_PARTPHI': self.material_tan_phi,
            'PV_PARTCOH': self.material_cohesie,
            'PV_VGWNAT_GEM': round(self.calc_vgwnat_gem, 3) if self.calc_vgwnat_gem is not None else None,
            'PV_VGWNAT_SD': round(self.calc_vgwnat_sd, 3) if self.calc_vgwnat_sd is not None else None,
            'PV_WATERGEHALTE_GEM': round(self.calc_watergehalte_gem,
                                         3) if self.calc_watergehalte_gem is not None else None,
            'PV_WATERGEHALTE_SD': round(self.calc_watergehalte_sd,
                                        3) if self.calc_watergehalte_sd is not None else None,
            'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        if file_path.exists():
            wb = load_workbook(file_path)
        else:
            template_path = Path(get_repo_root()) / "pv_tool" / "templates" / "Template_PVtool5_0.xlsx"
            wb = load_workbook(template_path)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            first_empty_row = ws.max_row + 1 if any(
                ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True)) else ws.max_row
        else:
            ws = wb.create_sheet(sheet_name)
            for col_idx, col_name in enumerate(expected_columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
            first_empty_row = 2

        for col_idx, col_name in enumerate(expected_columns, start=1):
            ws.cell(row=first_empty_row, column=col_idx, value=new_row.get(col_name, ""))

        wb.save(file_path)
        print(f"Resultaat toegevoegd aan template in tabblad '{sheet_name}'.")

    @property
    def save_total_to_excel(self):
        return lambda path: save_total_to_excel(self, path)

    @property
    def save_to_pdf(self):
        return lambda path: save_to_pdf(self, path)
